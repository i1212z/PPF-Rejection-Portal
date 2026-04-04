"""
Parse Due desk aging workbooks (.xlsx): company header, location sections (CLT / Kochi / TN),
and tabular rows with Particulars, SAFE, WARNING, DANGER, DOUBTFUL, TOTAL.

Supports:
- Single-row header: Particulars + zone columns + Total on one line.
- Two-row header (common in PPF exports): row1 Particulars / ZONE / TOTAL, row2 SAFE / WARNING / DANGER / DOUBTFUL.
- Merged cells (read_only mode may leave gaps; column indices still align with Excel columns).

Location rows: band titles like "Calicut Customers" (not company names containing PVT+LTD).
"""

from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

DATE_RANGE_RE = re.compile(
    r"\d{1,2}[-/][A-Za-z]{3}[-/]\d{2,4}\s+to\s+\d{1,2}[-/][A-Za-z]{3}[-/]\d{2,4}",
    re.I,
)


@dataclass
class ParsedAgingRow:
    location_group: str
    location_sort: int
    location_label: str
    particulars: str
    safe: float
    warning: float
    danger: float
    doubtful: float
    total: float


@dataclass
class ParsedWorkbook:
    company_title: str
    date_range_label: str
    rows: list[ParsedAgingRow]


def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _norm_key(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower())


def _parse_number(v: Any) -> float:
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    t = _cell_str(v).replace(",", "")
    if not t or t == "-":
        return 0.0
    try:
        return float(t)
    except ValueError:
        return 0.0


def _classify_location(text: str) -> tuple[str, int, str]:
    """Return (group_code, sort_key, display_label)."""
    raw = " ".join(text.split())
    low = raw.lower()
    if re.search(r"\bcalicut\b|\bclt\b", low):
        return ("CLT", 0, raw or "CLT / Calicut")
    if re.search(r"\bkochi\b|\bkottayam\b", low):
        return ("KOCHI", 1, raw or "Kochi")
    if re.search(r"\btamil\b|\btn\b|tamilnadu|tamil nadu|\bchennai\b|\bcoimbatore\b", low):
        return ("TN", 2, raw or "Tamil Nadu")
    return ("OTHER", 9, raw or "Other")


def _row_values(row: tuple[Any, ...]) -> list[str]:
    return [_cell_str(c) for c in row]


def _is_grand_total(particulars: str) -> bool:
    t = _norm_key(particulars)
    return "grand" in t and "total" in t


def _looks_like_company_header_line(low: str) -> bool:
    """Avoid treating '... PVT.LTD ... Calicut' title rows as location bands."""
    if re.search(r"\bpvt\b", low) and re.search(r"\b(ltd|limited)\b", low):
        return True
    if "international" in low and re.search(r"\b(farm|farms|pvt|ltd)\b", low):
        return True
    return False


def _has_geo_keyword(low: str) -> bool:
    return bool(
        re.search(
            r"\bcalicut\b|\bclt\b|\bkochi\b|\bkottayam\b|\btamil\b|\btn\b|tamilnadu|tamil nadu|\bchennai\b|\bcoimbatore\b",
            low,
        ),
    )


def _is_location_row(cells: list[str]) -> bool:
    """
    Section band only — not customer names like 'Beyondburg Inc Calicut' or 'Calicut Exhibition'.

    Treat as location when a geographic keyword appears together with a band word (Customers, …),
    or as a short Tamil Nadu / TN section title.
    """
    joined = " ".join(c for c in cells if c).strip()
    if not joined:
        return False
    low = joined.lower()
    if "particular" in low and "safe" in low:
        return False
    if not _has_geo_keyword(low):
        return False
    if _looks_like_company_header_line(low):
        return False
    if re.search(r"\bcustomers?\b|\bclients?\b|\bdesk\b|\bband\b|\bregister\b", low):
        return True
    if re.search(r"^tamil\s+nadu\b", low) or re.search(r"^tn\b", low):
        return True
    return False


def _joined_norm(cells: list[str]) -> str:
    return " ".join(_norm_key(c) for c in cells)


def _is_single_line_header_row(cells: list[str]) -> bool:
    """Particulars + at least one zone label on the same row."""
    j = _joined_norm(cells)
    if "particular" not in j:
        return False
    if "safe" not in j and "warning" not in j:
        return False
    return True


def _is_header_row_part1(cells: list[str]) -> bool:
    """
    First row of a two-row header: Particulars + ZONE + TOTAL (no SAFE/WARNING on this row).
    """
    j = _joined_norm(cells)
    if "particular" not in j:
        return False
    if "safe" in j or "warning" in j or "danger" in j or "doubtful" in j:
        return False
    return "zone" in j or "total" in j


def _is_zone_labels_row(cells: list[str]) -> bool:
    """Second row: SAFE, WARNING, DANGER, DOUBTFUL (no Particulars)."""
    j = _joined_norm(cells)
    if "particular" in j:
        return False
    hits = sum(1 for k in ("safe", "warning", "danger", "doubtful") if k in j)
    return hits >= 2


def _map_header_indices(cells: list[str]) -> dict[str, int]:
    """Map bucket keys to 0-based column index (single-row header)."""
    idx: dict[str, int] = {}
    for i, c in enumerate(cells):
        k = _norm_key(c)
        if not k:
            continue
        if "particular" in k:
            idx.setdefault("particulars", i)
        if k == "safe" or k.startswith("safe "):
            idx.setdefault("safe", i)
        if "warning" in k and "doubt" not in k:
            idx.setdefault("warning", i)
        if "danger" in k:
            idx.setdefault("danger", i)
        if "doubtful" in k:
            idx.setdefault("doubtful", i)
        if k == "total" or (k.startswith("total ") and "doubtful" not in k):
            idx.setdefault("total", i)
    return idx


def _map_two_row_header(row1: list[str], row2: list[str]) -> dict[str, int]:
    """Combine Particulars/TOTAL from row1 with zone columns from row2."""
    idx: dict[str, int] = {}
    for i, c in enumerate(row1):
        k = _norm_key(c)
        if not k:
            continue
        if "particular" in k:
            idx.setdefault("particulars", i)
        if k == "total" or (k.startswith("total ") and "doubtful" not in k):
            idx.setdefault("total", i)
    for i, c in enumerate(row2):
        k = _norm_key(c)
        if not k:
            continue
        if k == "safe" or k.startswith("safe "):
            idx.setdefault("safe", i)
        if "warning" in k and "doubt" not in k:
            idx.setdefault("warning", i)
        if "danger" in k:
            idx.setdefault("danger", i)
        if "doubtful" in k:
            idx.setdefault("doubtful", i)
    if "particulars" not in idx:
        idx["particulars"] = 0
    return idx


def parse_due_aging_xlsx(file_bytes: bytes) -> ParsedWorkbook:
    # read_only=False: slightly heavier but matches real column positions with merged cells better.
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=False)
    try:
        ws = wb.active
        title_guess = ""
        date_range = ""
        preamble_done = False

        current_group = "OTHER"
        current_sort = 9
        current_label = "General"

        col_map: dict[str, int] = {}
        in_table = False
        out_rows: list[ParsedAgingRow] = []
        pending_header_part1: list[str] | None = None

        for row in ws.iter_rows(values_only=True):
            cells = _row_values(row)
            if not any(x for x in cells):
                continue

            joined_display = " ".join(cells).strip()
            low_line = joined_display.lower()

            m_dates = DATE_RANGE_RE.search(joined_display)
            if m_dates:
                date_range = m_dates.group(0).strip()

            if pending_header_part1 is not None:
                if _is_zone_labels_row(cells):
                    preamble_done = True
                    col_map = _map_two_row_header(pending_header_part1, cells)
                    pending_header_part1 = None
                    in_table = True
                    continue
                pending_header_part1 = None

            if not preamble_done and not _is_location_row(cells) and not _is_single_line_header_row(cells) and not _is_header_row_part1(cells):
                if cells[0] and len(cells[0]) > 3:
                    if not title_guess or len(cells[0]) > len(title_guess):
                        title_guess = cells[0]
                low0 = low_line
                if "particular" not in low0 and not _looks_like_company_header_line(low0):
                    title_guess = title_guess or cells[0]

            if _is_location_row(cells):
                preamble_done = True
                grp, srt, lab = _classify_location(joined_display)
                current_group, current_sort, current_label = grp, srt, lab
                in_table = False
                col_map = {}
                pending_header_part1 = None
                continue

            if _is_single_line_header_row(cells):
                preamble_done = True
                col_map = _map_header_indices(cells)
                pending_header_part1 = None
                if "particulars" not in col_map:
                    col_map["particulars"] = 0
                in_table = True
                continue

            if _is_header_row_part1(cells):
                preamble_done = True
                pending_header_part1 = cells
                in_table = False
                col_map = {}
                continue

            if not in_table or not col_map:
                continue

            pi = col_map.get("particulars", 0)
            particulars = cells[pi] if pi < len(cells) else ""
            if not particulars or not particulars.strip():
                continue
            if _is_grand_total(particulars):
                continue

            def col(name: str) -> float:
                j = col_map.get(name)
                if j is None or j >= len(cells):
                    return 0.0
                return _parse_number(cells[j])

            s = col("safe")
            w = col("warning")
            dg = col("danger")
            db = col("doubtful")
            tot = col("total")
            if tot == 0.0 and (s or w or dg or db):
                tot = s + w + dg + db

            out_rows.append(
                ParsedAgingRow(
                    location_group=current_group,
                    location_sort=current_sort,
                    location_label=current_label,
                    particulars=particulars.strip(),
                    safe=s,
                    warning=w,
                    danger=dg,
                    doubtful=db,
                    total=tot,
                ),
            )

        if not title_guess:
            title_guess = ""

        return ParsedWorkbook(
            company_title=title_guess.strip(),
            date_range_label=date_range.strip(),
            rows=out_rows,
        )
    finally:
        wb.close()
